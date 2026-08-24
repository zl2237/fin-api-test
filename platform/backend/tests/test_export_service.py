"""export_service 单测：列表导出（Excel 简表 + JSON 全量）。

用 SimpleNamespace 构造 ORM 替身，验证 Excel 结构（sheet 名/表头/行数）、
JSON 结构（type/count/字段明细/节点配置合并）与 _fmt 的类型处理。
"""
import json
from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services import export_service as svc


def _field(key, ftype="string", required=False, default=None):
    return SimpleNamespace(
        key=key, label=f"{key}标签", field_type=ftype, required=required,
        default_value=default, remark=None, sort_order=0,
    )


def _api(aid, name="查询接口", method="POST", path="/api/x", group_id=1, fields=None):
    return SimpleNamespace(
        id=aid, name=name, code=f"api_{aid}", group_id=group_id,
        method=method, path=path, description="描述",
        sort_order=0, fields=fields or [],
        created_by_name="张三", updated_by_name="李四",
        created_at=datetime(2026, 1, 15, 10, 0, 0),
    )


def _nc(node_id, api_id=1):
    return SimpleNamespace(
        node_id=node_id, api_id=api_id,
        pre_process=[{"k": "v"}], post_extract=[{"as": "order_id"}],
        assertions=[{"type": "response_status_equals", "expected": 200}],
        wait_after_ms=100,
    )


def _case(cid, nodes=None, edges=None, configs=None, group_id=1):
    return SimpleNamespace(
        id=cid, name=f"用例{cid}", group_id=group_id, description="d",
        dag_config={"nodes": nodes or [], "edges": edges or []},
        node_configs=configs or [],
        created_by_name="张三", updated_by_name="李四",
        updated_at=datetime(2026, 1, 15, 11, 0, 0),
    )


GROUPS = {1: "冒烟组", 2: "订单组"}


class TestFmt:
    def test_none(self):
        assert svc._fmt(None) == ""

    def test_datetime(self):
        assert svc._fmt(datetime(2026, 1, 2, 3, 4, 5)) == "2026-01-02 03:04:05"

    def test_dict_json(self):
        assert json.loads(svc._fmt({"a": 1})) == {"a": 1}

    def test_other_to_str(self):
        assert svc._fmt(123) == "123"


class TestExportApisExcel:
    def test_structure(self):
        apis = [_api(1, fields=[_field("order_id", "int", True, "${bl_no}")]),
                _api(2, name="创建订单", method="GET", path="/api/y", group_id=2)]
        content = svc.export_apis_excel(apis, GROUPS)
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        assert ws.title == "接口列表"
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][:5] == ("ID", "名称", "分组", "方法", "路径")
        assert len(rows) == 3
        assert rows[1][1] == "查询接口" and rows[1][2] == "冒烟组" and rows[1][5] == 1
        assert rows[2][2] == "订单组"


class TestExportApisJson:
    def test_structure_and_fields(self):
        apis = [_api(1, fields=[_field("order_id", "int", True, "${bl_no}")])]
        data = json.loads(svc.export_apis_json(apis, GROUPS, "物流系统"))
        assert data["type"] == "apis" and data["count"] == 1 and data["project"] == "物流系统"
        item = data["items"][0]
        assert item["group"] == "冒烟组"
        assert item["fields"][0] == {
            "key": "order_id", "label": "order_id标签", "field_type": "int",
            "required": True, "default_value": "${bl_no}", "remark": None, "sort_order": 0,
        }

    def test_ungrouped_as_none(self):
        data = json.loads(svc.export_apis_json([_api(1, group_id=99)], GROUPS, "p"))
        assert data["items"][0]["group"] is None


class TestExportCasesExcel:
    def test_structure(self):
        cases = [_case(1, nodes=[{"id": "a"}, {"id": "b"}]), _case(2)]
        content = svc.export_cases_excel(cases, GROUPS)
        ws = load_workbook(BytesIO(content)).active
        assert ws.title == "用例列表"
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3
        assert rows[0][3] == "节点数"
        assert rows[1][3] == 2 and rows[2][3] == 0


class TestExportCasesJson:
    def test_node_configs_merged(self):
        """JSON 节点与 case_node_configs 按 node_id 合并，config 全量"""
        nodes = [
            {"id": "n1", "position": {"x": 1, "y": 2}, "data": {"label": "创建"}},
            {"id": "n2", "position": {"x": 3, "y": 4}, "data": {"label": "分发"}},
        ]
        edges = [{"id": "e1", "source": "n1", "target": "n2"}]
        cases = [_case(1, nodes=nodes, edges=edges, configs=[_nc("n1", api_id=7)])]
        data = json.loads(svc.export_cases_json(cases, GROUPS, "物流系统"))
        assert data["type"] == "cases" and data["count"] == 1
        item = data["items"][0]
        assert item["edges"] == edges
        n1, n2 = item["nodes"]
        assert n1["api_id"] == 7 and n1["config"]["post_extract"] == [{"as": "order_id"}]
        assert n1["config"]["wait_after_ms"] == 100
        # 无配置节点：config 为 None 不报错
        assert n2["config"] is None and n2["api_id"] is None
