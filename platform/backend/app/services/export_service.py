"""列表导出服务：Excel 简表（人看）+ JSON 全量（备份/迁移）。

口径：跟随列表页的后端筛选条件（project_id / created_by / updated_by）；
keyword 为前端本地过滤，不参与导出。Excel 列为摘要级；JSON 含字段、断言、节点配置全量。
"""
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(ws, ncols: int) -> None:
    """表头样式：加粗 + 浅灰底"""
    fill = PatternFill("solid", fgColor="F2F3F5")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def export_apis_excel(apis: list, group_names: dict[int, str]) -> bytes:
    """接口列表 Excel 简表。apis 为 ApiDefinition ORM 列表（已 fill_audit_names）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "接口列表"
    headers = ["ID", "名称", "分组", "方法", "路径", "字段数", "创建人", "更新人", "创建时间", "描述"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for a in apis:
        fields = a.fields or []
        ws.append([
            a.id, a.name, group_names.get(a.group_id, "") if a.group_id else "未分组",
            a.method, a.path, len(fields),
            getattr(a, "created_by_name", "") or "", getattr(a, "updated_by_name", "") or "",
            _fmt(a.created_at), a.description or "",
        ])
    _autosize(ws, [7, 26, 16, 8, 44, 8, 10, 10, 19, 28])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_apis_json(apis: list, group_names: dict[int, str], project_name: str) -> bytes:
    """接口全量 JSON：含字段明细，可后续做导入还原。"""
    data = {
        "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": "apis",
        "project": project_name,
        "count": len(apis),
        "items": [
            {
                "id": a.id, "name": a.name,
                "group": group_names.get(a.group_id) if a.group_id else None,
                "method": a.method, "path": a.path, "description": a.description,
                "sort_order": a.sort_order,
                "fields": [
                    {"key": f.key, "label": f.label, "field_type": f.field_type,
                     "required": f.required, "default_value": f.default_value,
                     "remark": f.remark, "sort_order": f.sort_order}
                    for f in (a.fields or [])
                ],
                "created_at": _fmt(a.created_at),
            }
            for a in apis
        ],
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


def export_cases_excel(cases: list, group_names: dict[int, str]) -> bytes:
    """用例列表 Excel 简表。cases 为 TestCase ORM 列表（已 fill_audit_names + node_configs 预载）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "用例列表"
    headers = ["ID", "名称", "分组", "节点数", "创建人", "更新人", "更新时间", "描述"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for c in cases:
        n = len((c.dag_config or {}).get("nodes", []))
        ws.append([
            c.id, c.name, group_names.get(c.group_id, "") if c.group_id else "未分组",
            n, getattr(c, "created_by_name", "") or "", getattr(c, "updated_by_name", "") or "",
            _fmt(c.updated_at), c.description or "",
        ])
    _autosize(ws, [7, 40, 16, 8, 10, 10, 19, 30])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_cases_json(cases: list, group_names: dict[int, str], project_name: str) -> bytes:
    """用例全量 JSON：含 DAG 结构与节点配置（断言/提取/前置处理），用于备份/迁移。"""
    items = []
    for c in cases:
        cfg_map = {nc.node_id: nc for nc in (c.node_configs or [])}
        nodes_out = []
        for n in (c.dag_config or {}).get("nodes", []):
            nc = cfg_map.get(n.get("id"))
            nodes_out.append({
                "id": n.get("id"),
                "label": (n.get("data") or {}).get("label"),
                "api_id": nc.api_id if nc else None,
                "position": n.get("position"),
                "config": {
                    "pre_process": (nc.pre_process or []) if nc else [],
                    "post_extract": (nc.post_extract or []) if nc else [],
                    "assertions": (nc.assertions or []) if nc else [],
                    "wait_after_ms": (nc.wait_after_ms or 0) if nc else 0,
                } if nc else None,
            })
        items.append({
            "id": c.id, "name": c.name,
            "group": group_names.get(c.group_id) if c.group_id else None,
            "description": c.description,
            "edges": (c.dag_config or {}).get("edges", []),
            "nodes": nodes_out,
            "updated_at": _fmt(c.updated_at),
        })
    data = {
        "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": "cases",
        "project": project_name,
        "count": len(items),
        "items": items,
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
