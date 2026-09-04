"""dataset_service：数据驱动测试的数据集服务层。

职责：
- 数据集 CRUD 校验与编排（列定义校验、引用保护、级联删行）
- 行操作（追加/改行/删行重排序）
- 导入解析（Excel/CSV → 列+行，周期 3 实现）
- 从用例生成（写死参数快照 + 节点配置快照）/ 复制 / 快照重新同步

概念定案（用例级 1:N）：
- 数据集归用例私有（case_id），用例间隔离，复用靠复制；
- 列中文名不落库，实时引用项目字段字典（FieldDictionary）；
- node_configs 为编排配置快照，执行时按 node_id 整块替换用例节点配置。

列名即变量名（方案定案 #4）：列 key 直接进执行变量池，因此校验比普通
命名更严——点号撞嵌套路径语法、${} 撞表达式占位符、空格撞变量引用。
"""
import io
import json
import re
from copy import deepcopy
from typing import Any

from sqlalchemy.orm import Session

from .. import crud, models
from ..engine.topo import topo_order
from .body_builder import parse_field_value

# 列 key 合法字符：字母数字下划线（不允许点/空格/${}，理由见模块注释）
_COL_KEY_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

_VALID_COL_TYPES = {"string", "int", "bool", "array", "object", "file"}


def _validate_columns(columns: list) -> None:
    """列定义校验：非空、key 唯一且合法、type 在白名单；label 剥离（中文名实时引用字段字典）"""
    if not columns:
        raise ValueError("数据集至少需要一列")
    seen = set()
    for col in columns:
        key = str(col.get("key") or "").strip()
        if not _COL_KEY_RE.match(key):
            raise ValueError(f"列名 {key!r} 不合法：仅允许字母/数字/下划线且不以数字开头（列名即变量名，点号/空格/表达式符会与表达式语法冲突）")
        if key in seen:
            raise ValueError(f"列名重复：{key}（列名即变量名，重复会相互覆盖）")
        seen.add(key)
        col["key"] = key
        col.pop("label", None)  # 列中文名实时引用项目字段字典，不落库
        ctype = col.get("type") or "string"
        if ctype not in _VALID_COL_TYPES:
            raise ValueError(f"列 {key} 类型 {ctype!r} 不支持：{'/'.join(sorted(_VALID_COL_TYPES))}")
        col["type"] = ctype


def _validate_row_data(columns: list, data: dict) -> None:
    """行数据校验：key 必须都在列定义内（缺列允许=置空语义）"""
    keys = {c["key"] for c in columns}
    unknown = set(data.keys()) - keys
    if unknown:
        raise ValueError(f"行数据含未定义的列：{'、'.join(sorted(unknown))}")


# ============ 数据集 CRUD ============

def create_dataset(db: Session, project_id: int, name: str, columns: list,
                   user_id: int, description: str = None,
                   rows_data: list = None, case_id: int = None,
                   node_configs: list = None) -> models.DataSet:
    """建数据集；rows_data/node_configs 可选（创建即带行/带快照的原子写入路径）"""
    columns = [dict(c) for c in columns]
    _validate_columns(columns)
    obj = models.DataSet(project_id=project_id, case_id=case_id, name=name, description=description,
                         columns=columns, node_configs=node_configs or [],
                         created_by=user_id, updated_by=user_id)
    db.add(obj)
    if rows_data:
        db.flush()  # 先拿 id，行数据外键依赖
        for data in rows_data:
            _validate_row_data(columns, data)
        for i, data in enumerate(rows_data, start=1):
            db.add(models.DataSetRow(dataset_id=obj.id, row_index=i, data=data))
    db.commit()
    db.refresh(obj)
    return obj


def get_dataset(db: Session, dataset_id: int):
    obj = crud.get_dataset(db, dataset_id)
    if not obj:
        raise ValueError(f"数据集不存在: {dataset_id}")
    return obj


def update_dataset(db: Session, dataset_id: int, name: str = None,
                   description: str = None, columns: list = None) -> models.DataSet:
    """更新数据集。改列定义时校验现有行数据不含被删的列（防行悬空）。"""
    obj = get_dataset(db, dataset_id)
    if name is not None:
        obj.name = name
    if description is not None:
        obj.description = description
    if columns is not None:
        new_cols = [dict(c) for c in columns]
        _validate_columns(new_cols)
        new_keys = {c["key"] for c in new_cols}
        for row in (obj.rows or []):
            used = set((row.data or {}).keys()) - new_keys
            if used:
                raise ValueError(f"行 #{row.row_index} 的行数据仍使用已删列：{'、'.join(sorted(used))}，请先清理行数据")
        obj.columns = new_cols
    db.commit()
    db.refresh(obj)
    return obj


def delete_dataset(db: Session, dataset_id: int) -> None:
    """删除数据集：被用例绑定时拒绝；成功则级联删行（ORM cascade）"""
    obj = get_dataset(db, dataset_id)
    n = crud.count_cases_bound_to_dataset(db, dataset_id)
    if n:
        raise ValueError(f"数据集被 {n} 个用例绑定，请先在用例中解绑后再删除")
    db.query(models.DataSetRow).filter(models.DataSetRow.dataset_id == dataset_id).delete()
    db.delete(obj)
    db.commit()


# ============ 行操作 ============

def add_row(db: Session, dataset_id: int, data: dict) -> models.DataSetRow:
    """追加行：row_index = 现有最大行序 + 1"""
    ds = get_dataset(db, dataset_id)
    _validate_row_data(ds.columns, data)
    next_index = max((r.row_index for r in (ds.rows or [])), default=0) + 1
    row = models.DataSetRow(dataset_id=dataset_id, row_index=next_index, data=data)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def update_row(db: Session, dataset_id: int, row_id: int, data: dict) -> models.DataSetRow:
    ds = get_dataset(db, dataset_id)
    row = crud.get_row(db, row_id)
    if not row or row.dataset_id != dataset_id:
        raise ValueError(f"行不存在: {row_id}")
    _validate_row_data(ds.columns, data)
    row.data = data
    db.commit()
    db.refresh(row)
    return row


def copy_row(db: Session, dataset_id: int, row_id: int) -> models.DataSetRow:
    """复制行：原行数据深拷贝追加为新行（row_index 顺延），复用 add_row 的校验与编号逻辑"""
    row = crud.get_row(db, row_id)
    if not row or row.dataset_id != dataset_id:
        raise ValueError(f"行不存在: {row_id}")
    return add_row(db, dataset_id, data=dict(row.data or {}))


def delete_row(db: Session, dataset_id: int, row_id: int) -> None:
    """删行并重排行序保持连续（1..n）"""
    get_dataset(db, dataset_id)  # 存在性校验（不存在时 raise）
    row = crud.get_row(db, row_id)
    if not row or row.dataset_id != dataset_id:
        raise ValueError(f"行不存在: {row_id}")
    db.delete(row)
    db.commit()
    # 重排：删除后重新查询（不能复用内存列表，可能含已删对象），按 row_index 升序重新编号
    remaining = crud.list_rows(db, dataset_id)
    for i, r in enumerate(remaining, start=1):
        r.row_index = i
    db.commit()


def replace_rows(db: Session, dataset_id: int, rows_data: list) -> list:
    """批量保存（表格整页保存语义）：先全量校验再整体替换，row_index 从 1 连续编号。

    原子性：任何一行校验失败则整批拒绝，不动库。
    """
    ds = get_dataset(db, dataset_id)
    for data in rows_data:
        _validate_row_data(ds.columns, data)
    db.query(models.DataSetRow).filter(models.DataSetRow.dataset_id == dataset_id).delete()
    rows = [models.DataSetRow(dataset_id=dataset_id, row_index=i, data=data)
            for i, data in enumerate(rows_data, start=1)]
    for row in rows:
        db.add(row)
    db.commit()
    return rows


def clear_rows(db: Session, dataset_id: int) -> None:
    """全清行（不删数据集本身，列定义保留）"""
    get_dataset(db, dataset_id)
    db.query(models.DataSetRow).filter(models.DataSetRow.dataset_id == dataset_id).delete()
    db.commit()


# ============ 执行展开（方案定案 #3：每行一条执行记录） ============

def plan_case_expansion(db: Session, case, dataset_id=None, row_ids=None) -> list:
    """按用例绑定生成执行展开计划。

    - 未绑定 dataset_id → [{dataset_id: None, row: None}]（单条普通执行，行为与现状一致）
    - 绑定 N 行 → N 条，每条含行快照 {row_index, data, label}（label=首列值，展示用）
    - 绑定但 0 行 → ValueError（先录入数据再执行）
    - dataset_id 传入时临时覆盖用例绑定（执行面板换数据集，不改绑定本身）
    - row_ids 传入时只执行选中行（单行手动执行=逐条通知的来源）
    - origins：列快照原值 {key: 生成时值}，执行时据此做快照保真过滤（同名异值列
      只作用于配置值与原值一致的节点；手工列/旧数据集无 origin → 不过滤）
    """
    effective = dataset_id if dataset_id is not None else getattr(case, "dataset_id", None)
    if not effective:
        return [{"dataset_id": None, "row": None, "overrides": None, "origins": None}]
    ds = crud.get_dataset(db, effective)
    if not ds:
        raise ValueError(f"用例绑定的数据集不存在: {effective}")
    if ds.case_id != case.id:
        raise ValueError("数据集不属于该用例（数据集按用例隔离，请在用例自己的数据集中选择）")
    rows = crud.list_rows(db, effective)
    if row_ids is not None:
        wanted = set(row_ids)
        unknown = wanted - {r.id for r in rows}
        if unknown:
            raise ValueError(f"数据行不存在: {'、'.join(str(u) for u in sorted(unknown))}")
        rows = [r for r in rows if r.id in wanted]
    if not rows:
        raise ValueError("数据集无数据行，请先录入数据再执行")
    first_key = (ds.columns or [{}])[0].get("key") if ds.columns else None
    # 节点配置快照映射：命中 node_id 整块替换用例编排（数据集=场景包语义）
    overrides = {c["node_id"]: c for c in (ds.node_configs or []) if c.get("node_id")} or None
    origins = {c["key"]: c["origin"] for c in (ds.columns or [])
               if isinstance(c, dict) and "origin" in c} or None
    return [{
        "dataset_id": effective,
        "row": {
            "row_index": r.row_index,
            "data": dict(r.data or {}),
            "label": str((r.data or {}).get(first_key)) if first_key else str(r.row_index),
        },
        "overrides": overrides,
        "origins": origins,
    } for r in rows]


# ============ 用例绑定校验（方案定案 #5：用例级绑定，同项目） ============

def validate_binding(db: Session, case, dataset_id) -> None:
    """用例绑定数据集前的校验：None=解绑直过；须存在且为本用例名下（用例间隔离，复用靠复制）"""
    if dataset_id is None:
        return
    ds = crud.get_dataset(db, dataset_id)
    if not ds:
        raise ValueError(f"数据集不存在: {dataset_id}")
    if ds.case_id != case.id:
        raise ValueError("不能绑定其他用例的数据集（数据集按用例隔离，复用请复制）")


# ============ 从用例生成数据集（写死参数快照） ============

def _infer_col_type(value) -> str:
    """列类型按值推断（bool 先于 int：bool 是 int 子类）"""
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return "string"


def _topo_node_ids(dag: dict) -> list:
    """dag 节点的拓扑执行序（engine.topo_order 唯一实现，此处口径：环/断链节点排末尾）。

    收集口径需按执行序取值：同名异值列取业务链路源头节点（执行序首个）的值，
    而非 dag 数组顺序（数组可能是用户拖拽后的任意顺序，源头节点不一定排最前）。
    环/断链节点按引擎同口径排到末尾（引擎也不执行它们，只是收集时仍扫一遍兜底）。
    """
    order, leftover = topo_order(dag)
    return order + leftover


def collect_case_params(case, node_configs: list, apis_by_id: dict) -> dict:
    """扫描用例全部节点，收集可参数化的"写死"请求参数（纯函数，不触 db）。

    收集口径与三级取值优先级一致（数据集 = 除动态绑定外的所有字段集合）：
    - API 字段默认值：顶层 key（无点号）、非空、不含 ${}（上游提取注入属动态，不收）；
      file 字段同样成列（值是文件中心文件 ID，列 type=file，行编辑器渲染文件选择器，
      执行时经 pop_file_fields_from_body 剥离走 multipart，链路无缝）；
      GET 接口例外：空值 query 参数（可选过滤条件）也成列——执行时行值为空不覆盖
      （apply_row_overrides 跳过空值），用户填值才启用该参数，天然参数化语义；
      但 key 不符 _COL_KEY_RE（如 curl 导入的 search_time[date] 带方括号）跳过：
      列名即变量名，无法引用也无法行值覆盖（stats.invalid 计数提示）
    - 前置处理 set_field/add_field：path 顶层、value 不含 ${}；
      同节点同 key 时 set_field 覆盖默认值（执行顺序在组装之后，最终生效值为准）
    - set_field/add_field 顶层 path 的 value 含 ${}（运行时动态注入，如 [${audit_id}]）
      → 该 key 整体剔除（含早前节点收集的默认值）：列存在会触发"列名==path 行值优先"
      拦截表达式求值，动态注入字段不能成为数据列
    - 跨节点同名 → 合并一列（一列统一覆盖所有同名节点）；同名异值取执行序首个
      （业务链路源头）节点的值，并把该值记为列 origin（快照原值）；
      stats.conflicts 记录异值字段仅作提示（并集口径：除动态绑定外所有字段都成列）
    - 列 origin（生成时该 key 的源头值）：执行时快照保真——行值只作用于
      "节点自身配置值 == origin"的节点，异值节点保留自身配置（见 filter_row_vars_for_node），
      保证生成的数据集原样执行与原用例行为一致

    返回 {columns, row, stats}；row = 1 行原值快照（同名异值列取源头节点值，改值即参数化）。
    全部不可提取时返回空结果（columns=[]），由 generate_dataset_from_case 抛用户可读错误。
    """
    values: dict = {}      # key -> 首个最终生效值（列顺序 = 拓扑执行序）
    key_types: dict = {}   # key -> 字段类型（file 需透传到列定义，行编辑器据此渲染文件选择器）
    dynamic_keys = set()   # 被 ${} 动态注入的顶层 key：运行时表达式生效，不可作数据列
    conflicts = []
    stats = {"nodes": 0, "columns": 0, "conflicts": conflicts, "dynamic": 0,
             "nested": 0, "empty": 0, "invalid": 0}

    cfg_by_node = {c.node_id: c for c in node_configs}
    for node_id in _topo_node_ids(getattr(case, "dag_config", None)):
        cfg = cfg_by_node.get(node_id)
        if not cfg or not cfg.api_id:
            continue
        api = apis_by_id.get(cfg.api_id)
        if not api:
            continue
        stats["nodes"] += 1

        # key -> field_type：pre_process 路径的 file 字段值同为文件 ID，需识别类型
        api_field_types = {f.key: (f.field_type or "string") for f in (getattr(api, "fields", None) or []) if f.key}
        # GET 的 query 参数：空值（可选过滤条件）也成列，用户填值才启用（执行层空行值不覆盖）
        is_get = str(getattr(api, "method", "") or "").upper() == "GET"

        node_vals: dict = {}  # 本节点最终生效值：默认值 → set_field 覆盖
        node_types: dict = {}  # 本节点 key -> 字段类型
        for f in getattr(api, "fields", None) or []:
            if not f.key:
                continue
            if "." in f.key:
                stats["nested"] += 1
                continue
            # 列名即变量名（_COL_KEY_RE）：curl 导入的 GET query 原始参数名可能含
            # 方括号（如 search_time[date]），无法作变量名也无法行值覆盖 → 跳过保真执行
            if not _COL_KEY_RE.match(f.key):
                stats["invalid"] += 1
                continue
            raw = f.default_value
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                if not is_get:
                    stats["empty"] += 1
                    continue
                node_vals[f.key] = ""
                node_types[f.key] = f.field_type or "string"
                continue
            if isinstance(raw, str) and "${" in raw:
                stats["dynamic"] += 1
                continue
            node_vals[f.key] = parse_field_value(raw, f.field_type or "string")
            node_types[f.key] = f.field_type or "string"
        for act in cfg.pre_process or []:
            if act.get("type") not in ("set_field", "add_field"):
                continue
            path = act.get("path") or ""
            val = act.get("value")
            if not path:
                continue  # 空行占位（前端表格留空）非有效动作
            if "." in path:
                stats["nested"] += 1
                continue
            if not _COL_KEY_RE.match(path):
                stats["invalid"] += 1
                continue
            if isinstance(val, str) and "${" in val:
                # 动态注入（如 [${audit_id}]）：运行时由表达式求值，不是数据列；
                # 剔除本节点与早前节点收集的同名值，防止列拦截表达式
                stats["dynamic"] += 1
                dynamic_keys.add(path)
                node_vals.pop(path, None)
                values.pop(path, None)
                continue
            node_vals[path] = val
            node_types[path] = api_field_types.get(path, "string")

        for k, v in node_vals.items():
            if k in dynamic_keys:
                continue
            if k not in values:
                values[k] = v
                key_types[k] = node_types.get(k, "string")
            elif values[k] != v:
                # 同名异值：并集口径仍成一列（数据集=除动态绑定外所有字段的集合），
                # 取执行序源头节点值；conflicts 仅提示异值。
                # 执行时 origin 保真：异值节点不应用行值（见 filter_row_vars_for_node）
                conflicts.append({"key": k, "values": [values[k], v]})

    if not values:
        # 全部不可提取（动态/嵌套/空/冲突）：返回空结果，由 generate_dataset_from_case 抛用户可读错误
        return {"columns": [], "row": {}, "stats": stats}

    # origin = 生成时该 key 的源头值（与 row 初值相同）：执行时快照保真的比对基准；
    # file 列类型来自字段定义（值是文件 ID，按值推断只会得到 string，行编辑器无法识别）
    columns = [
        {"key": k, "type": "file" if key_types.get(k) == "file" else _infer_col_type(v), "origin": v}
        for k, v in values.items()
    ]
    stats["columns"] = len(columns)
    return {"columns": columns, "row": dict(values), "stats": stats}


def filter_row_vars_for_node(row_vars: dict | None, origins: dict | None,
                             api, pre_process: list | None) -> dict | None:
    """快照保真过滤：数据集行值按节点过滤后应用（纯函数，执行层调用）。

    同名异值列（跨节点配置值不同，如 main_ids '3,1' vs ',3,1,'）一张列无法同时
    驱动所有节点——按行值是否被用户编辑分两种语义：

    1. 行值 == origin（生成后未编辑的快照值）：只作用于"节点自身配置值 == origin"
       的节点，异值节点保留自身配置。防的是生成时被下游真实订单节点污染的快照
       盖掉源头节点配置（teu 56→3 案例），保证原样执行与原用例行为一致。
    2. 行值 != origin（用户在数据集里编辑过该单元格）：明确的覆盖意图，
       作用于全部节点——即使某节点配置值与 origin 异值（如接口默认值是从真实
       订单拷贝的"带 order_id 版本"），也按行值覆盖（supplier 整体替换场景）。

    - origins 为 None/空（手工列、旧数据集无 origin）→ 不过滤，行为与现状一致
    - 节点对 key 的自身配置值 = pre_process 顶层 set_field/add_field 字面量 ?? API 字段
      默认值（parse_field_value 解析后，与生成口径一致）
    - 默认值为空的字段记空串哨兵（≠ 任何非空 origin）→ 排除：空值节点的字段
      不被其他节点贡献的列值盖掉（生成只收非空值，保真口径空 ≠ 非空）
    - pre_process 对该 key 是动态注入（值含 ${}）→ 该 key 排除（动态绑定不在数据集范围）
    - 节点没有该 key（请求体无此字段）→ 保留（apply_row_overrides 只覆盖已存在字段，无副作用）
    """
    if not row_vars or not origins:
        return row_vars
    effective: dict = {}
    excluded = set()  # 动态注入 key：不在数据集覆盖范围（生成侧已剔除，此处双保险）
    for f in getattr(api, "fields", None) or []:
        if not f.key or "." in f.key or f.field_type == "file":
            continue
        raw = f.default_value
        if isinstance(raw, str) and "${" in raw:
            continue  # 动态默认值：apply_row_overrides 本就跳过 ${} 值
        effective[f.key] = "" if raw is None or (isinstance(raw, str) and not raw.strip()) \
            else parse_field_value(raw, f.field_type or "string")
    for act in pre_process or []:
        if act.get("type") not in ("set_field", "add_field"):
            continue
        path = act.get("path") or ""
        if not path or "." in path:
            continue
        val = act.get("value")
        if isinstance(val, str) and "${" in val:
            excluded.add(path)  # 动态注入：该 key 不在数据集覆盖范围
        else:
            effective[path] = val
    result: dict = {}
    for k, v in row_vars.items():
        if k in excluded:
            continue
        if k not in origins or v != origins[k]:
            # 无 origin（手工列/旧数据集）或用户编辑过该单元格 → 无条件覆盖
            result[k] = v
        elif k not in effective or effective[k] == origins[k]:
            # 未编辑的快照值：仅当节点配置与 origin 一致才应用（防污染）
            result[k] = v
    return result


def snapshot_node_configs(case, node_configs: list) -> list:
    """把用例当前节点编排（前置/后置/断言/等待/接口绑定）快照为数据集配置（纯函数）。

    只快照 dag_config 中存在的节点（执行时按 node_id 匹配，对不上自然回落用例配置）。
    """
    node_ids = {n.get("id") for n in (getattr(case, "dag_config", None) or {}).get("nodes", [])}
    return [{
        "node_id": c.node_id,
        "api_id": c.api_id,
        "pre_process": deepcopy(c.pre_process or []),
        "post_extract": deepcopy(c.post_extract or []),
        "assertions": deepcopy(c.assertions or []),
        "wait_after_ms": c.wait_after_ms or 0,
    } for c in node_configs if c.node_id in node_ids]


def generate_dataset_from_case(db: Session, case_id: int, name: str = None,
                               user_id: int = None) -> tuple:
    """从用例生成数据集：写死参数各成一列 + 1 行原值快照 + 节点配置快照，返回 (dataset, stats)。

    生成的数据集归属该用例（case_id，1:N 隔离）。
    """
    case = crud.get_testcase(db, case_id)
    if not case:
        raise ValueError(f"用例不存在: {case_id}")
    # DataSet.name 上限 100：自定义名超长直接报错；默认名截断用例名部分保后缀
    suffix = "-参数集"
    if name:
        if len(name) > 100:
            raise ValueError("名称最多 100 字符")
    else:
        name = f"{case.name}{suffix}"
        if len(name) > 100:
            name = case.name[:100 - len(suffix) - 1] + f"…{suffix}"
    cfgs = (db.query(models.CaseNodeConfig)
            .filter(models.CaseNodeConfig.case_id == case_id).all())
    api_ids = {c.api_id for c in cfgs if c.api_id}
    apis = (db.query(models.ApiDefinition)
            .filter(models.ApiDefinition.id.in_(api_ids)).all()) if api_ids else []
    out = collect_case_params(case, cfgs, {a.id: a for a in apis})
    if not out["columns"]:
        raise ValueError("该用例没有可提取的写死请求参数（字段全为 ${} 动态注入 / 嵌套路径 / 空值 / 跨节点同名异值），无法生成数据集")
    ds = create_dataset(
        db, project_id=case.project_id, case_id=case.id,
        name=name,
        columns=out["columns"], user_id=user_id,
        description=f"从用例「{case.name}」的写死请求参数生成（{out['stats']['columns']} 列，含 1 行原值快照）",
        rows_data=[out["row"]],
        node_configs=snapshot_node_configs(case, cfgs),
    )
    return ds, out["stats"]


def copy_dataset(db: Session, dataset_id: int, name: str = None, user_id: int = None) -> models.DataSet:
    """复制数据集：列/全部行/节点配置快照全量深拷贝，归属同一用例（隔离语义下的复用方式）。

    命名：默认「原名-副本」；原名已带 -副本 后缀时递增编号（场景A-副本 → 场景A-副本2），
    与该用例名下已有名冲突时继续递增。
    """
    src = get_dataset(db, dataset_id)
    base = name or src.name
    if not name:
        if base.endswith("-副本") or "-副本" in base:
            stem, _, num = base.rpartition("-副本")
            try:
                n = int(num) if num else 1
            except ValueError:
                stem, n = base, 1
            base = f"{stem}-副本{n + 1}"
        else:
            base = f"{base}-副本"
        # 同用例下重名递增
        dup = (db.query(models.DataSet)
               .filter(models.DataSet.case_id == src.case_id, models.DataSet.name == base)
               .first())
        while dup:
            stem, _, num = base.rpartition("-副本")
            n = int(num) + 1 if num.isdigit() else 2
            base = f"{stem}-副本{n}"
            dup = (db.query(models.DataSet)
                   .filter(models.DataSet.case_id == src.case_id, models.DataSet.name == base)
                   .first())
        if len(base) > 100:
            raise ValueError("名称最多 100 字符")
    elif len(name) > 100:
        raise ValueError("名称最多 100 字符")
    return create_dataset(
        db, project_id=src.project_id, case_id=src.case_id,
        name=base, user_id=user_id,
        description=src.description,
        columns=deepcopy(src.columns or []),
        node_configs=deepcopy(src.node_configs or []),
        rows_data=[deepcopy(r.data or {}) for r in (src.rows or [])],
    )


def resync_node_configs(db: Session, dataset_id: int) -> int:
    """把数据集归属用例的当前节点编排重新快照进数据集（列/行数据不动），返回快照节点数。"""
    ds = get_dataset(db, dataset_id)
    case = crud.get_testcase(db, ds.case_id)
    if not case:
        raise ValueError(f"归属用例不存在: {ds.case_id}")
    cfgs = (db.query(models.CaseNodeConfig)
            .filter(models.CaseNodeConfig.case_id == ds.case_id).all())
    ds.node_configs = snapshot_node_configs(case, cfgs)
    db.commit()
    return len(ds.node_configs)


def sync_case_datasets(db: Session, case_id: int) -> int:
    """保存用例后自动同步：把该用例绑定的全部数据集 node_configs 一次性重快照（列/行不动）。

    与手动 resync_node_configs 同口径（snapshot_node_configs），按用例批量执行；
    各数据集持有快照的独立深拷贝，互不影响。返回同步的数据集数
    （未绑定数据集或用例已不存在返回 0，不视为错误）。
    """
    datasets = (db.query(models.DataSet)
                .filter(models.DataSet.case_id == case_id).all())
    if not datasets:
        return 0
    case = crud.get_testcase(db, case_id)
    if not case:
        return 0
    cfgs = (db.query(models.CaseNodeConfig)
            .filter(models.CaseNodeConfig.case_id == case_id).all())
    snapshot = snapshot_node_configs(case, cfgs)
    for ds in datasets:
        ds.node_configs = deepcopy(snapshot)
    db.commit()
    return len(datasets)


# 编排字段中文名（drift 报告用）
_DRIFT_FIELD_NAMES = {
    "api_id": "接口绑定",
    "pre_process": "前置处理",
    "post_extract": "后置提取",
    "assertions": "断言",
    "wait_after_ms": "等待时长",
}


def config_drift(db: Session, dataset_id: int) -> dict:
    """比对数据集节点配置快照与归属用例当前编排，返回差异清单（执行前提示快照过期）。

    只比快照与用例两侧都存在的节点：仅快照有（用例已删节点，执行不再触发）或
    仅有用例（快照缺位，执行自然回落用例当前配置）都不影响执行行为。
    字段级 diff 与 snapshot_node_configs 同口径（None/空列表归一）。
    """
    ds = get_dataset(db, dataset_id)
    case = crud.get_testcase(db, ds.case_id)
    if not case:
        raise ValueError(f"归属用例不存在: {ds.case_id}")
    # 当前编排按快照同口径归一（只取 dag 中存在的节点）
    cfgs = (db.query(models.CaseNodeConfig)
            .filter(models.CaseNodeConfig.case_id == ds.case_id).all())
    current = {c["node_id"]: c for c in snapshot_node_configs(case, cfgs)}
    snapshot = {c["node_id"]: c for c in (ds.node_configs or []) if c.get("node_id")}
    labels = {n.get("id"): (n.get("label") or n.get("id"))
              for n in (getattr(case, "dag_config", None) or {}).get("nodes", [])
              if isinstance(n, dict)}
    nodes = []
    for nid in sorted(set(snapshot) & set(current)):
        s, c = snapshot[nid], current[nid]
        changes = []
        for field, cname in _DRIFT_FIELD_NAMES.items():
            if s.get(field) != c.get(field):
                if field in ("pre_process", "post_extract", "assertions"):
                    changes.append(f"{cname}：{len(s.get(field) or [])} 条 → {len(c.get(field) or [])} 条")
                else:
                    changes.append(f"{cname}：{s.get(field)} → {c.get(field)}")
        if changes:
            nodes.append({"node_id": nid, "label": labels.get(nid, nid), "changes": changes})
    return {"stale": bool(nodes), "nodes": nodes}


# ============ 导入解析（纯函数，不触 db） ============

def _coerce_cell(value, col_type: str | None):
    """导入单元格按列类型还原：object/array 列的字符串值尝试 JSON 解析（失败保留原样）。

    与导出侧（export_rows_excel 序列化为 JSON 字符串）对偶，保证 Excel 往返后类型不丢
    （历史缺陷：object 列导入后以字符串落库，行值覆盖会把字符串塞进请求体）。
    """
    if col_type in ("object", "array") and isinstance(value, str) and value.strip():
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return value
    return value


def parse_import_file(filename: str, content: bytes, columns: list) -> tuple:
    """Excel/CSV → 行数据。首行表头按列 key 匹配。

    返回 (rows_data, warnings)：
    - 多余列（文件有、columns 没有）忽略 + 告警
    - 缺失列（columns 有、表头没有）置空 + 告警
    - 值不做类型转换：xlsx 保留原始类型（int/bool），CSV 全字符串，
      类型语义由 body_builder 按列 type 强转（方案定案 #4）
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        header, raw_rows = _parse_csv(content)
    elif name.endswith(".xlsx"):
        header, raw_rows = _parse_xlsx(content)
    else:
        raise ValueError("导入仅支持 .xlsx / .csv 格式")

    if not header:
        raise ValueError("导入文件为空或无表头行")
    if not raw_rows:
        raise ValueError("导入文件无数据行（只有表头）")

    warnings = []
    col_keys = [c["key"] for c in columns]
    col_types = {c["key"]: c.get("type") for c in columns}
    header_set = {h.strip() for h in header}
    extra = [h for h in header if h.strip() not in col_keys]
    if extra:
        warnings.append(f"文件中的列 {'、'.join(extra)} 不在数据集列定义内，已忽略")
    missing = [k for k in col_keys if k not in header_set]
    if missing:
        warnings.append(f"数据集列 {'、'.join(missing)} 在文件表头中缺失，对应值已置空")

    key_by_header = {h.strip(): h.strip() for h in header if h.strip() in col_keys}
    rows = []
    for raw in raw_rows:
        row_data = {k: "" for k in col_keys}  # 缺失列置空
        for i, cell in enumerate(raw):
            key = key_by_header.get(str(header[i]).strip()) if i < len(header) else None
            if key is not None:
                row_data[key] = "" if cell is None else cell
        # object/array 列的字符串值还原为原生 JSON（往返类型不丢）
        rows.append({k: _coerce_cell(v, col_types.get(k)) for k, v in row_data.items()})
    return rows, warnings


def _parse_csv(content: bytes):
    """CSV：UTF-8（含 BOM 兼容）解码，逗号分隔"""
    import csv as _csv
    text = content.decode("utf-8-sig")  # sig 自动剥离 BOM
    reader = _csv.reader(io.StringIO(text))
    matrix = [row for row in reader if any(str(c).strip() for c in row)]
    if not matrix:
        return [], []
    return [str(c).strip() for c in matrix[0]], matrix[1:]


def _parse_xlsx(content: bytes):
    """xlsx：openpyxl 读首张表，None 单元格保留为 None（置空语义在组装层处理）"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    matrix = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            matrix.append(list(row))
    wb.close()
    if not matrix:
        return [], []
    return [str(c).strip() if c is not None else "" for c in matrix[0]], matrix[1:]


# ============ 导出与覆盖合并（与导入对偶，数据集间字段流转） ============

def export_rows_excel(db: Session, dataset_id: int) -> tuple[bytes, str, int]:
    """数据集行导出 xlsx：首行表头=列 key，其后每行一行数据。

    object/array 值序列化为 JSON 字符串（导入侧 _coerce_cell 还原，往返类型不丢）。
    返回 (文件内容, 数据集名, 行数)；文件名由 router 组装（需时间戳）。
    """
    ds = get_dataset(db, dataset_id)
    rows = crud.list_rows(db, dataset_id)  # 已按 row_index 排序
    if not rows:
        raise ValueError("数据集无数据行，请先录入或导入数据再导出")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "rows"
    keys = [c["key"] for c in (ds.columns or [])]
    ws.append(keys)
    for r in rows:
        data = r.data or {}
        ws.append([_cell_out(data.get(k)) for k in keys])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), ds.name, len(rows)


def _cell_out(v) -> Any:
    """导出单元格序列化：None→空串；dict/list→JSON 字符串；其余原样。"""
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def _node_param_keys(cfg, api) -> set:
    """单节点可参数化列（collect_case_params 的单节点口径，含 dynamic 剔除）。

    cfg 为节点配置快照 dict（node_configs 元素）。
    - API 字段默认值：顶层、非 file、非空、不含 ${}
    - pre_process set_field/add_field：path 顶层、value 非动态
    - 动态注入（value 含 ${}）的 key 整体剔除（与生成口径一致）
    """
    keys: set = set()
    for f in getattr(api, "fields", None) or []:
        if not f.key or "." in f.key or f.field_type == "file":
            continue
        raw = f.default_value
        if raw is None or (isinstance(raw, str) and (not raw.strip() or "${" in raw)):
            continue
        keys.add(f.key)
    for act in (cfg.get("pre_process") or []):
        if act.get("type") not in ("set_field", "add_field"):
            continue
        path = act.get("path") or ""
        val = act.get("value")
        if not path or "." in path:
            continue
        if isinstance(val, str) and "${" in val:
            keys.discard(path)
            continue
        keys.add(path)
    return keys


def compare_datasets(db: Session, target_id: int, source_id: int) -> dict:
    """对比两个数据集的节点配置快照：按 api_id 配对相同节点，算出可覆盖列。

    可覆盖列 = 节点参数化列 ∩ 源数据集列 ∩ 目标数据集列（三方交集：
    列不在数据集里则行数据里没有值可刷，或刷了也不生效）。
    无相同节点 → common_nodes=[]（调用方据此提示"没有覆盖的必要"）。
    """
    if target_id == source_id:
        raise ValueError("源数据集与目标数据集不能相同")
    t = get_dataset(db, target_id)
    s = get_dataset(db, source_id)
    t_keys = {c["key"] for c in (t.columns or [])}
    s_keys = {c["key"] for c in (s.columns or [])}

    t_by_api: dict = {}
    for c in (t.node_configs or []):
        if c.get("api_id"):
            t_by_api.setdefault(c["api_id"], []).append(c)
    common = []
    seen = set()
    for c in (s.node_configs or []):
        aid = c.get("api_id")
        if not aid or aid not in t_by_api or aid in seen:
            continue
        seen.add(aid)
        api = db.get(models.ApiDefinition, aid)
        if not api:
            continue
        keys: set = set()
        for sc in [c, *t_by_api[aid]]:
            keys |= _node_param_keys(sc, api)
        columns = sorted(keys & t_keys & s_keys)
        if not columns:
            continue
        common.append({
            "api_id": aid,
            "api_name": api.name,
            "target_nodes": [x.get("node_id") for x in t_by_api[aid]],
            "source_nodes": [c.get("node_id")],
            "columns": columns,
        })
    src_rows = crud.list_rows(db, source_id)
    return {
        "source": {"id": s.id, "name": s.name, "case_id": s.case_id, "rows": len(src_rows),
                   "row_labels": [str(r.row_index) for r in src_rows]},
        "common_nodes": common,
        "columns_total": len({k for n in common for k in n["columns"]}),
    }


def merge_from_dataset(db: Session, target_id: int, source_id: int,
                       api_ids: list | None = None, source_row_index: int = 1) -> dict:
    """覆盖合并：源数据集指定行的"相同节点涉及列"值刷到目标数据集全部行。

    - api_ids 不传 = 全部相同节点；传则只刷这些节点的列（与 compare 结果对齐）
    - 源行空值（None/""）跳过：与行值覆盖语义一致（空=未配置，不覆盖不置空）
    - 目标独有列自然保留；目标行数为 0 时报错（无可作用对象）
    """
    cmp_data = compare_datasets(db, target_id, source_id)
    nodes = cmp_data["common_nodes"]
    if not nodes:
        raise ValueError("两个数据集没有相同节点，无覆盖的必要")
    if api_ids is not None:
        wanted = set(api_ids)
        unknown = wanted - {n["api_id"] for n in nodes}
        if unknown:
            raise ValueError(f"接口不在相同节点列表内: {'、'.join(str(u) for u in sorted(unknown))}")
        nodes = [n for n in nodes if n["api_id"] in wanted]
    columns = {k for n in nodes for k in n["columns"]}
    if not columns:
        raise ValueError("所选节点没有两侧数据集共有的可覆盖列")

    src_row = (db.query(models.DataSetRow)
               .filter(models.DataSetRow.dataset_id == source_id,
                       models.DataSetRow.row_index == source_row_index).first())
    if not src_row:
        raise ValueError(f"源数据集不存在第 {source_row_index} 行")
    overrides = {k: v for k, v in (src_row.data or {}).items()
                 if k in columns and v is not None and v != ""}
    if not overrides:
        raise ValueError("源行在所选节点列上均为空值，没有可覆盖内容")

    target_rows = crud.list_rows(db, target_id)
    if not target_rows:
        raise ValueError("目标数据集无数据行，覆盖合并无可作用对象")
    for r in target_rows:
        r.data = {**(r.data or {}), **overrides}
    db.commit()
    return {"rows": len(target_rows), "columns": len(overrides), "keys": sorted(overrides)}
